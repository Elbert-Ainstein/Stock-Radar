#!/usr/bin/env python3
"""
model_export.py — Export a price-target model to Excel with LIVE formulas.

This build is ALIGNED TO THE INSTITUTIONAL-GRADE TARGET ENGINE (target_engine.py):
  - 5-year forecast with margin ramp
  - Terminal value at end of Year 3 using EV/EBITDA AND EV/(FCF-SBC)
  - 3 INDEPENDENT scenarios (Down / Base / Up) with their own drivers
  - Discount from NTM+3 back to PV at scenario-specific WACC

Produces a 6-sheet workbook:
  1. Cover        — ticker, current price, target range, warnings
  2. Assumptions  — the driver inputs (blue) that every other sheet references
  3. P&L Summary  — historical + 5-year forecast revenue / EBITDA / FCF-SBC
  4. Income Stmt  — full historicals from finance_data.py + forecast rows
  5. Cash Flow    — historical cash flow lines + forecast FCF
  6. Valuation    — 3-scenario EV/EBITDA + EV/(FCF-SBC) blended, WACC-discounted

Every forecast cell is a FORMULA that reads from Assumptions — the user can
open the file and tweak any input to see the target move live. No hardcoded
forecast values anywhere.

Color coding follows GS sell-side convention:
    BLUE text    = hardcoded input (drivers, historicals)
    BLACK text   = in-sheet formula
    GREEN text   = cross-sheet link
    YELLOW fill  = editable input cell
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from finance_data import FinancialData, fetch_financials, EarningsFetchError
from target_engine import (
    build_target,
    TargetResult,
    DEFAULT_DRIVERS,
    DRIVER_META,
    SCENARIO_OFFSETS,
    VALUATION_YEAR,
    DISCOUNT_YEARS,
    MARGIN_RAMP_YEARS,
    FORECAST_YEARS,
)


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------
FONT = "Arial"
BLUE = Font(name=FONT, size=10, color="0000FF")
BLACK = Font(name=FONT, size=10, color="000000")
GREEN = Font(name=FONT, size=10, color="008000")
BOLD_BLACK = Font(name=FONT, size=10, color="000000", bold=True)
BOLD_BLUE = Font(name=FONT, size=10, color="0000FF", bold=True)
BOLD_WHITE = Font(name=FONT, size=11, color="FFFFFF", bold=True)
TITLE = Font(name=FONT, size=14, color="000000", bold=True)
ITALIC = Font(name=FONT, size=9, italic=True, color="666666")

HEADER_FILL = PatternFill("solid", start_color="1F3864")
SUBHEADER_FILL = PatternFill("solid", start_color="D9E1F2")
INPUT_FILL = PatternFill("solid", start_color="FFFFCC")
TOTAL_FILL = PatternFill("solid", start_color="E2E2E2")
BASE_COL_FILL = PatternFill("solid", start_color="EAF3FF")

CENTER = Alignment(horizontal="center", vertical="center")
RIGHT = Alignment(horizontal="right")
LEFT = Alignment(horizontal="left")

THIN = Side(border_style="thin", color="999999")
BOX = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

FMT_MONEY_MM = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'
FMT_PRICE = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
FMT_PCT = "0.0%;(0.0%);-"
FMT_MULT = "0.0\\x"
FMT_SHARES = "#,##0.0"

# Scenario → column mapping for Valuation tab
SCEN_COLS = {"downside": "C", "base": "D", "upside": "E"}


# ---------------------------------------------------------------------------
# Main export
# ---------------------------------------------------------------------------
def export_model(
    fin: FinancialData,
    target: TargetResult,
    out_path: str | Path,
) -> Path:
    wb = Workbook()
    wb.remove(wb.active)

    _build_cover(wb, fin, target)
    _build_assumptions(wb, fin, target)
    _build_pnl_summary(wb, fin, target)
    _build_income_stmt(wb, fin, target)
    _build_cash_flow(wb, fin, target)
    _build_valuation(wb, fin, target)

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


# ---------------------------------------------------------------------------
# Sheet: Cover
# ---------------------------------------------------------------------------
def _build_cover(wb: Workbook, fin: FinancialData, t: TargetResult) -> None:
    ws = wb.create_sheet("Cover")
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22

    ws["B2"] = f"{fin.ticker} — {fin.name}"
    ws["B2"].font = TITLE
    ws["B3"] = f"Sector: {fin.sector}  |  Currency: {fin.currency}"
    ws["B3"].font = Font(name=FONT, size=10, italic=True, color="555555")
    ws["B4"] = f"Data source: {fin.source}   |   Fetched: {fin.fetched_at[:19]}"
    ws["B4"].font = Font(name=FONT, size=9, italic=True, color="777777")

    r = 6
    ws.cell(row=r, column=2, value="Price target range").font = BOLD_BLACK
    ws.cell(row=r, column=2).fill = SUBHEADER_FILL
    r += 1
    for lbl, val, fmt in [
        ("Current price", t.current_price, FMT_PRICE),
        ("Downside target", t.low, FMT_PRICE),
        ("Base target", t.base, FMT_PRICE),
        ("Upside target", t.high, FMT_PRICE),
        ("Upside to base", t.upside_base_pct, FMT_PCT),
        ("Valuation horizon", t.terminal_year, None),
    ]:
        ws.cell(row=r, column=2, value=lbl).font = BLACK
        c = ws.cell(row=r, column=3, value=val)
        c.font = BLUE
        if fmt:
            c.number_format = fmt
        c.alignment = RIGHT
        r += 1

    r += 1
    ws.cell(row=r, column=2, value="Capitalization").font = BOLD_BLACK
    ws.cell(row=r, column=2).fill = SUBHEADER_FILL
    r += 1
    for lbl, val, fmt in [
        ("Market cap ($mm)", (fin.market_cap or 0) / 1e6, FMT_MONEY_MM),
        ("Diluted shares (mm)", (fin.shares_diluted or 0) / 1e6, FMT_SHARES),
        ("Net debt ($mm)", (fin.net_debt or 0) / 1e6, FMT_MONEY_MM),
    ]:
        ws.cell(row=r, column=2, value=lbl).font = BLACK
        c = ws.cell(row=r, column=3, value=val)
        c.font = BLUE
        c.number_format = fmt
        c.alignment = RIGHT
        r += 1

    r += 1
    ws.cell(row=r, column=2, value="TTM actuals").font = BOLD_BLACK
    ws.cell(row=r, column=2).fill = SUBHEADER_FILL
    r += 1
    for lbl, val in [
        ("TTM revenue ($mm)", (fin.ttm_revenue() or 0) / 1e6),
        ("TTM operating income ($mm)", (fin.ttm_operating_income() or 0) / 1e6),
        ("TTM EBITDA ($mm)", (fin.ttm_ebitda() or 0) / 1e6),
        ("TTM free cash flow ($mm)", (fin.ttm_fcf() or 0) / 1e6),
        ("TTM FCF − SBC ($mm)", (t.ttm_fcf_sbc or 0) / 1e6),
    ]:
        ws.cell(row=r, column=2, value=lbl).font = BLACK
        c = ws.cell(row=r, column=3, value=val)
        c.font = BLUE
        c.number_format = FMT_MONEY_MM
        c.alignment = RIGHT
        r += 1

    if t.warnings:
        r += 1
        ws.cell(row=r, column=2, value="Warnings").font = BOLD_BLACK
        ws.cell(row=r, column=2).fill = PatternFill("solid", start_color="FFE8CC")
        r += 1
        for w in t.warnings:
            ws.cell(row=r, column=2, value=f"• {w}").font = Font(
                name=FONT, size=9, color="884400"
            )
            ws.merge_cells(start_row=r, end_row=r, start_column=2, end_column=5)
            r += 1

    r += 2
    ws.cell(row=r, column=2, value="Note: Forecast cells are LIVE FORMULAS referencing 'Assumptions'.").font = ITALIC
    r += 1
    ws.cell(row=r, column=2, value="Change any blue input to re-price. Scenarios derive from base drivers via fixed offsets.").font = ITALIC


# ---------------------------------------------------------------------------
# Sheet: Assumptions
# ---------------------------------------------------------------------------
def _build_assumptions(wb: Workbook, fin: FinancialData, t: TargetResult) -> dict[str, str]:
    """Single source of driver truth. Other sheets reference cells from this tab."""
    ws = wb.create_sheet("Assumptions")
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 50

    ws["B2"] = "Driver Assumptions"
    ws["B2"].font = TITLE
    ws["B3"] = "Blue = base-case driver (user-editable). Black = formula-derived scenarios."
    ws["B3"].font = ITALIC

    for i, h in enumerate(["Driver", "Value", "Description"]):
        c = ws.cell(row=5, column=2 + i, value=h)
        c.font = BOLD_WHITE
        c.fill = HEADER_FILL
        c.alignment = CENTER

    refs: dict[str, str] = {}
    r = 6
    # Base-case drivers
    for key in DEFAULT_DRIVERS:
        meta = DRIVER_META.get(key, {})
        label = meta.get("label", key)
        fmt = meta.get("format", "pct")
        ws.cell(row=r, column=2, value=label).font = BLACK
        c = ws.cell(row=r, column=3, value=t.drivers.get(key, DEFAULT_DRIVERS[key]))
        c.font = BLUE
        c.fill = INPUT_FILL
        c.alignment = RIGHT
        if fmt == "pct":
            c.number_format = FMT_PCT
        elif fmt == "mult":
            c.number_format = FMT_MULT
        else:
            c.number_format = "0.000"
        desc = _default_desc(key)
        dc = ws.cell(row=r, column=4, value=desc)
        dc.font = ITALIC
        refs[key] = f"Assumptions!$C${r}"
        r += 1

    r += 1
    ws.cell(row=r, column=2, value="— Base period actuals —").font = BOLD_BLACK
    ws.cell(row=r, column=2).fill = SUBHEADER_FILL
    r += 1
    anchors: list[tuple[str, str, float, str]] = [
        ("ttm_revenue",  "TTM revenue ($mm)",           (fin.ttm_revenue() or 0) / 1e6,     FMT_MONEY_MM),
        ("ttm_opinc",    "TTM operating income ($mm)",  (fin.ttm_operating_income() or 0) / 1e6, FMT_MONEY_MM),
        ("ttm_ebitda",   "TTM EBITDA ($mm)",            (fin.ttm_ebitda() or 0) / 1e6,      FMT_MONEY_MM),
        ("ttm_fcf",      "TTM free cash flow ($mm)",    (fin.ttm_fcf() or 0) / 1e6,         FMT_MONEY_MM),
        ("ttm_fcf_sbc",  "TTM FCF − SBC ($mm)",         (t.ttm_fcf_sbc or 0) / 1e6,         FMT_MONEY_MM),
        ("shares_0",     "Diluted shares, base (mm)",   (fin.shares_diluted or 0) / 1e6,    FMT_SHARES),
        ("net_debt",     "Net debt ($mm)",              (fin.net_debt or 0) / 1e6,          FMT_MONEY_MM),
        ("price_0",      "Current price ($)",           fin.price or 0.0,                   FMT_PRICE),
    ]
    for key, lbl, val, fmt in anchors:
        ws.cell(row=r, column=2, value=lbl).font = BLACK
        c = ws.cell(row=r, column=3, value=val)
        c.font = BLUE
        c.fill = INPUT_FILL
        c.number_format = fmt
        c.alignment = RIGHT
        refs[key] = f"Assumptions!$C${r}"
        r += 1

    # Scenario offset block (computed from SCENARIO_OFFSETS constants)
    r += 1
    ws.cell(row=r, column=2, value="— Scenario offsets (calibrated vs institutional models) —").font = BOLD_BLACK
    ws.cell(row=r, column=2).fill = SUBHEADER_FILL
    r += 1
    for i, h in enumerate(["Offset", "Downside", "Base", "Upside"]):
        c = ws.cell(row=r, column=2 + i, value=h)
        c.font = BOLD_WHITE
        c.fill = HEADER_FILL
        c.alignment = CENTER
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    r += 1

    scenario_keys = [
        ("rev_growth_y1_mult",    "Rev growth Y1 (vs base)",    FMT_MULT),
        ("rev_growth_terminal_mult", "Rev growth terminal (vs base)", FMT_MULT),
        ("ebitda_margin_delta",   "EBITDA margin delta (pp)",   FMT_PCT),
        ("fcf_sbc_margin_delta",  "FCF-SBC margin delta (pp)",  FMT_PCT),
        ("ev_ebitda_multiple_mult", "EV/EBITDA mult (vs base)",   FMT_MULT),
        ("ev_fcf_sbc_multiple_mult", "EV/(FCF-SBC) mult (vs base)", FMT_MULT),
    ]
    scen_refs: dict[str, dict[str, str]] = {s: {} for s in ("downside", "base", "upside")}
    for key, label, fmt in scenario_keys:
        ws.cell(row=r, column=2, value=label).font = BLACK
        for j, scen in enumerate(("downside", "base", "upside")):
            v = SCENARIO_OFFSETS[scen][key]
            c = ws.cell(row=r, column=3 + j, value=v)
            c.font = BLUE
            c.fill = INPUT_FILL
            c.number_format = fmt
            c.alignment = RIGHT
            scen_refs[scen][key] = f"Assumptions!${get_column_letter(3 + j)}${r}"
        r += 1

    wb._target_refs = refs                    # type: ignore[attr-defined]
    wb._scen_refs = scen_refs                 # type: ignore[attr-defined]
    return refs


def _default_desc(key: str) -> str:
    return {
        "rev_growth_y1": "YoY revenue growth, forecast Year 1",
        "rev_growth_terminal": "YoY revenue growth, Year 5 (linear decay Y1→Y5)",
        "ebitda_margin_target": "Terminal (Y3) EBITDA margin, ramped from TTM",
        "fcf_sbc_margin_target": "Terminal (Y3) FCF-minus-SBC margin",
        "ev_ebitda_multiple": "Exit EV / Y3 EBITDA multiple (base scenario)",
        "ev_fcf_sbc_multiple": "Exit EV / Y3 (FCF-SBC) multiple (base scenario)",
        "discount_rate": "WACC for discounting terminal EV to present (base scenario)",
        "tax_rate": "Effective tax rate on pretax income",
        "da_pct_rev": "D&A as % of revenue (for EBITDA bridge)",
        "sbc_pct_rev": "Stock-based comp as % of revenue",
        "share_change_pct": "Annual net dilution (+) or buyback (−)",
    }.get(key, "")


# ---------------------------------------------------------------------------
# Sheet: P&L Summary (historical + 5-year forecast, formula-driven forecast)
# ---------------------------------------------------------------------------
def _build_pnl_summary(wb: Workbook, fin: FinancialData, t: TargetResult) -> None:
    ws = wb.create_sheet("P&L Summary")
    refs = wb._target_refs                    # type: ignore[attr-defined]

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 32
    for i in range(15):
        ws.column_dimensions[get_column_letter(3 + i)].width = 13

    ws["B2"] = "P&L Summary — historical actuals + 5-year base-case forecast"
    ws["B2"].font = TITLE
    ws["B3"] = "($mm — forecast cells are formulas referencing Assumptions)"
    ws["B3"].font = ITALIC

    ann = [p for p in fin.annual_income if p.get("Total Revenue")][-4:]
    hist_labels = [p["period"] for p in ann]
    hist_revs = [p.get("Total Revenue") or 0 for p in ann]
    hist_oi = [p.get("Operating Income") or 0 for p in ann]
    hist_ni = [p.get("Net Income") or 0 for p in ann]

    # Index annual cashflow by period label so we can match D&A / SBC / FCF to the
    # corresponding income-statement period (labels don't always align 1-for-1).
    cf_by_period: dict[str, dict] = {p["period"]: p for p in fin.annual_cashflow}

    def _cf_val(period: str, key: str) -> float:
        p = cf_by_period.get(period)
        if not p:
            return 0.0
        v = p.get(key)
        return float(v) if v is not None else 0.0

    # Historical EBITDA = Operating Income + D&A (D&A from cash flow statement)
    hist_ebitda = [
        (hist_oi[i] + _cf_val(hist_labels[i], "Depreciation And Amortization"))
        for i in range(len(hist_labels))
    ]
    # Historical FCF − SBC = Free Cash Flow − Stock-Based Comp (both from cash flow)
    hist_fcf_sbc = [
        (_cf_val(hist_labels[i], "Free Cash Flow")
         - _cf_val(hist_labels[i], "Stock Based Compensation"))
        for i in range(len(hist_labels))
    ]

    forecast_labels = [p.period for p in t.forecast_annual]  # 5 years
    all_labels = hist_labels + forecast_labels

    header_row = 5
    ws.cell(row=header_row, column=2, value="($mm)").font = BOLD_WHITE
    ws.cell(row=header_row, column=2).fill = HEADER_FILL
    for i, lbl in enumerate(all_labels):
        c = ws.cell(row=header_row, column=3 + i, value=lbl)
        c.font = BOLD_WHITE
        c.fill = HEADER_FILL
        c.alignment = CENTER

    n_hist = len(hist_labels)
    n_fwd = len(forecast_labels)

    # Row numbers we'll use for cross-sheet references
    REV_ROW = 6
    REVGR_ROW = 7
    EBITDA_MGN_ROW = 8
    EBITDA_ROW = 9
    FCFSBC_MGN_ROW = 10
    FCFSBC_ROW = 11
    OI_ROW = 12
    NI_ROW = 13

    # --- Revenue ---
    ws.cell(row=REV_ROW, column=2, value="Revenue").font = BOLD_BLACK
    for i, v in enumerate(hist_revs):
        c = ws.cell(row=REV_ROW, column=3 + i, value=v / 1e6)
        c.font = BLUE
        c.number_format = FMT_MONEY_MM

    # Forecast revenue: Y1 anchors to TTM (not last annual) so the P&L Summary
    # stays consistent with the engine and Valuation tab, which both compound
    # from TTM. Anchoring to FY_last would diverge whenever TTM ≠ FY_last
    # (e.g., MU's Aug-ending fiscal year or any ticker where the latest reported
    # quarter sits outside the most recent annual period).
    # Subsequent years chain off the prior forecast year.
    for j in range(n_fwd):
        col = get_column_letter(3 + n_hist + j)
        # Growth rate for this year: g1 + (g_term - g1) * (y-1)/(FORECAST_YEARS-1)
        yfrac = j / (FORECAST_YEARS - 1) if FORECAST_YEARS > 1 else 0
        g_formula = (
            f"({refs['rev_growth_y1']}+({refs['rev_growth_terminal']}-{refs['rev_growth_y1']})*{yfrac})"
        )
        if j == 0:
            formula = f"={refs['ttm_revenue']}*(1+{g_formula})"
        else:
            prev_col = get_column_letter(3 + n_hist + j - 1)
            formula = f"={prev_col}{REV_ROW}*(1+{g_formula})"
        c = ws.cell(row=REV_ROW, column=3 + n_hist + j, value=formula)
        c.font = BLACK
        c.number_format = FMT_MONEY_MM

    # --- Revenue YoY% ---
    ws.cell(row=REVGR_ROW, column=2, value="Revenue YoY %").font = BLACK
    for i in range(1, n_hist + n_fwd):
        col = get_column_letter(3 + i)
        prev = get_column_letter(3 + i - 1)
        c = ws.cell(row=REVGR_ROW, column=3 + i, value=f"=IFERROR({col}{REV_ROW}/{prev}{REV_ROW}-1,0)")
        c.font = BLACK
        c.number_format = FMT_PCT

    # --- EBITDA margin % (historical computed as formula = EBITDA / Revenue) ---
    ws.cell(row=EBITDA_MGN_ROW, column=2, value="EBITDA margin %").font = BLACK
    for i in range(n_hist):
        col = get_column_letter(3 + i)
        c = ws.cell(row=EBITDA_MGN_ROW, column=3 + i,
                    value=f"=IFERROR({col}{EBITDA_ROW}/{col}{REV_ROW},0)")
        c.font = BLACK
        c.number_format = FMT_PCT
    for j in range(n_fwd):
        y = j + 1
        ramp = min(1.0, y / MARGIN_RAMP_YEARS)
        # margin = ttm_mgn + (target - ttm_mgn) * ramp
        formula = (
            f"=({refs['ttm_ebitda']}/{refs['ttm_revenue']})"
            f"+({refs['ebitda_margin_target']}-{refs['ttm_ebitda']}/{refs['ttm_revenue']})*{ramp}"
        )
        c = ws.cell(row=EBITDA_MGN_ROW, column=3 + n_hist + j, value=formula)
        c.font = BLACK
        c.number_format = FMT_PCT

    # --- EBITDA (historical = OI + D&A as hardcoded input, forecast = margin × rev) ---
    ws.cell(row=EBITDA_ROW, column=2, value="EBITDA").font = BOLD_BLACK
    for i, v in enumerate(hist_ebitda):
        c = ws.cell(row=EBITDA_ROW, column=3 + i, value=v / 1e6)
        c.font = BLUE
        c.number_format = FMT_MONEY_MM
    for j in range(n_fwd):
        col = get_column_letter(3 + n_hist + j)
        formula = f"={col}{REV_ROW}*{col}{EBITDA_MGN_ROW}"
        c = ws.cell(row=EBITDA_ROW, column=3 + n_hist + j, value=formula)
        c.font = BLACK
        c.number_format = FMT_MONEY_MM

    # --- FCF-SBC margin % (historical as formula, forecast ramps to target) ---
    ws.cell(row=FCFSBC_MGN_ROW, column=2, value="FCF − SBC margin %").font = BLACK
    for i in range(n_hist):
        col = get_column_letter(3 + i)
        c = ws.cell(row=FCFSBC_MGN_ROW, column=3 + i,
                    value=f"=IFERROR({col}{FCFSBC_ROW}/{col}{REV_ROW},0)")
        c.font = BLACK
        c.number_format = FMT_PCT
    for j in range(n_fwd):
        y = j + 1
        ramp = min(1.0, y / MARGIN_RAMP_YEARS)
        formula = (
            f"=({refs['ttm_fcf_sbc']}/{refs['ttm_revenue']})"
            f"+({refs['fcf_sbc_margin_target']}-{refs['ttm_fcf_sbc']}/{refs['ttm_revenue']})*{ramp}"
        )
        c = ws.cell(row=FCFSBC_MGN_ROW, column=3 + n_hist + j, value=formula)
        c.font = BLACK
        c.number_format = FMT_PCT

    # --- FCF − SBC (historical = FCF − SBC from cash flow, forecast = margin × rev) ---
    ws.cell(row=FCFSBC_ROW, column=2, value="FCF − SBC").font = BOLD_BLACK
    for i, v in enumerate(hist_fcf_sbc):
        c = ws.cell(row=FCFSBC_ROW, column=3 + i, value=v / 1e6)
        c.font = BLUE
        c.number_format = FMT_MONEY_MM
    for j in range(n_fwd):
        col = get_column_letter(3 + n_hist + j)
        formula = f"={col}{REV_ROW}*{col}{FCFSBC_MGN_ROW}"
        c = ws.cell(row=FCFSBC_ROW, column=3 + n_hist + j, value=formula)
        c.font = BLACK
        c.number_format = FMT_MONEY_MM

    # --- Operating income (historicals, forecast derived) ---
    ws.cell(row=OI_ROW, column=2, value="Operating income").font = BLACK
    for i, v in enumerate(hist_oi):
        c = ws.cell(row=OI_ROW, column=3 + i, value=v / 1e6)
        c.font = BLUE
        c.number_format = FMT_MONEY_MM
    # Forecast: OI = Rev × (ttm_op_mgn + MAX(0, (ebitda_target − da_pct_rev) − ttm_op_mgn) × ramp)
    # Matches engine's _forecast_annual exactly: op_target = max(ttm_op_mgn, ebitda_target − d&a_pct)
    # and m_op = ttm_op_mgn + (op_target − ttm_op_mgn) × ramp.
    for j in range(n_fwd):
        y = j + 1
        ramp = min(1.0, y / MARGIN_RAMP_YEARS)
        col = get_column_letter(3 + n_hist + j)
        ttm_mgn = f"({refs['ttm_opinc']}/{refs['ttm_revenue']})"
        op_target = f"({refs['ebitda_margin_target']}-{refs['da_pct_rev']})"
        formula = (
            f"={col}{REV_ROW}*({ttm_mgn}+MAX(0,{op_target}-{ttm_mgn})*{ramp})"
        )
        c = ws.cell(row=OI_ROW, column=3 + n_hist + j, value=formula)
        c.font = BLACK
        c.number_format = FMT_MONEY_MM

    # --- Net income (historical from income stmt, forecast tax-effected from OI) ---
    ws.cell(row=NI_ROW, column=2, value="Net income").font = BLACK
    for i, v in enumerate(hist_ni):
        c = ws.cell(row=NI_ROW, column=3 + i, value=v / 1e6)
        c.font = BLUE
        c.number_format = FMT_MONEY_MM
    for j in range(n_fwd):
        col = get_column_letter(3 + n_hist + j)
        formula = f"={col}{OI_ROW}*(1-{refs['tax_rate']})"
        c = ws.cell(row=NI_ROW, column=3 + n_hist + j, value=formula)
        c.font = BLACK
        c.number_format = FMT_MONEY_MM

    # Stash cross-sheet coordinates so Valuation tab can link (e.g., Y3 EBITDA)
    wb._pnl_coords = {                         # type: ignore[attr-defined]
        "first_forecast_col": 3 + n_hist,
        "rev_row": REV_ROW,
        "ebitda_row": EBITDA_ROW,
        "fcfsbc_row": FCFSBC_ROW,
        "oi_row": OI_ROW,
        "ni_row": NI_ROW,
    }


# ---------------------------------------------------------------------------
# Sheet: Income Statement — last 5 quarters + 4 forecast quarters
# ---------------------------------------------------------------------------
def _build_income_stmt(wb: Workbook, fin: FinancialData, t: TargetResult) -> None:
    ws = wb.create_sheet("Income Stmt")
    refs = wb._target_refs                     # type: ignore[attr-defined]

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 36

    periods = fin.quarterly_income[-5:]
    q_labels = [p["period"] for p in periods]
    forecast_q = t.forecast_quarterly[:4]
    all_labels = q_labels + [p.period for p in forecast_q]
    for i in range(len(all_labels)):
        ws.column_dimensions[get_column_letter(3 + i)].width = 13

    ws["B2"] = "Income Statement ($mm) — quarterly"
    ws["B2"].font = TITLE

    header_row = 5
    ws.cell(row=header_row, column=2, value="Line item").font = BOLD_WHITE
    ws.cell(row=header_row, column=2).fill = HEADER_FILL
    for i, lbl in enumerate(all_labels):
        c = ws.cell(row=header_row, column=3 + i, value=lbl)
        c.font = BOLD_WHITE
        c.fill = HEADER_FILL
        c.alignment = CENTER

    is_rows = [
        ("Total Revenue", "Total Revenue", True),
        ("Cost Of Revenue", "Cost Of Revenue", False),
        ("Gross Profit", "Gross Profit", True),
        ("R&D", "Research And Development", False),
        ("SG&A", "Selling General And Administration", False),
        ("Operating Income", "Operating Income", True),
        ("Interest Expense", "Interest Expense", False),
        ("Pretax Income", "Pretax Income", False),
        ("Tax Provision", "Tax Provision", False),
        ("Net Income", "Net Income", True),
    ]

    n_hist = len(q_labels)
    REV_ROW_Q = 6
    COGS_ROW_Q = 7
    GP_ROW_Q = 8
    OI_ROW_Q = 11
    NI_ROW_Q = 15
    r = 6
    for lbl, key, bold in is_rows:
        ws.cell(row=r, column=2, value=lbl).font = BOLD_BLACK if bold else BLACK
        for i, p in enumerate(periods):
            v = p.get(key)
            c = ws.cell(row=r, column=3 + i, value=(v / 1e6) if v is not None else None)
            c.font = BLUE
            c.number_format = FMT_MONEY_MM
        r += 1

    # Last historical column letter — used as anchor for margin carries into forecast
    last_hist_col_q = get_column_letter(3 + n_hist - 1) if n_hist > 0 else None

    # Forecast quarterly revenue = prior-year same quarter × (1 + g1)
    # Forecast OI = Rev × ramp(TTM op margin → op_target)
    for j, fp in enumerate(forecast_q):
        py_idx = n_hist - 4 + j
        py_col = get_column_letter(3 + py_idx) if py_idx >= 0 else get_column_letter(3)
        col = get_column_letter(3 + n_hist + j)

        c = ws.cell(row=REV_ROW_Q, column=3 + n_hist + j,
                    value=f"={py_col}{REV_ROW_Q}*(1+{refs['rev_growth_y1']})")
        c.font = BLACK
        c.number_format = FMT_MONEY_MM

        # Forecast Gross Profit = Revenue × (last historical gross margin)
        # and Cost of Revenue = Revenue − Gross Profit, so the waterfall and
        # Gross margin % row stay coherent for forecast periods.
        if last_hist_col_q:
            c = ws.cell(
                row=GP_ROW_Q, column=3 + n_hist + j,
                value=f"={col}{REV_ROW_Q}*IFERROR({last_hist_col_q}{GP_ROW_Q}/{last_hist_col_q}{REV_ROW_Q},0)",
            )
            c.font = BLACK
            c.number_format = FMT_MONEY_MM
            c = ws.cell(
                row=COGS_ROW_Q, column=3 + n_hist + j,
                value=f"={col}{REV_ROW_Q}-{col}{GP_ROW_Q}",
            )
            c.font = BLACK
            c.number_format = FMT_MONEY_MM

        # Op margin ramp across 4 quarters: fraction = (j+1)/4 scaled vs one-year advance
        ramp_frac = (j + 1) / 4
        c = ws.cell(
            row=OI_ROW_Q, column=3 + n_hist + j,
            value=(
                f"={col}{REV_ROW_Q}*({refs['ttm_opinc']}/{refs['ttm_revenue']}"
                f"+(({refs['ebitda_margin_target']}-{refs['da_pct_rev']})"
                f"-{refs['ttm_opinc']}/{refs['ttm_revenue']})*{ramp_frac})"
            ),
        )
        c.font = BLACK
        c.number_format = FMT_MONEY_MM

        c = ws.cell(row=NI_ROW_Q, column=3 + n_hist + j,
                    value=f"={col}{OI_ROW_Q}*(1-{refs['tax_rate']})")
        c.font = BLACK
        c.number_format = FMT_MONEY_MM

    # Margin rows
    r = 17
    for lbl, num_row in [("Gross margin %", 8), ("Operating margin %", OI_ROW_Q), ("Net margin %", NI_ROW_Q)]:
        ws.cell(row=r, column=2, value=lbl).font = BLACK
        for i in range(n_hist + len(forecast_q)):
            col = get_column_letter(3 + i)
            c = ws.cell(row=r, column=3 + i, value=f"=IFERROR({col}{num_row}/{col}{REV_ROW_Q},0)")
            c.font = BLACK
            c.number_format = FMT_PCT
        r += 1


# ---------------------------------------------------------------------------
# Sheet: Cash Flow
# ---------------------------------------------------------------------------
def _build_cash_flow(wb: Workbook, fin: FinancialData, t: TargetResult) -> None:
    ws = wb.create_sheet("Cash Flow")
    refs = wb._target_refs                     # type: ignore[attr-defined]

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 32
    periods = fin.quarterly_cashflow[-5:]
    q_labels = [p["period"] for p in periods]
    forecast_q = t.forecast_quarterly[:4]
    all_labels = q_labels + [p.period for p in forecast_q]
    for i in range(len(all_labels)):
        ws.column_dimensions[get_column_letter(3 + i)].width = 13

    ws["B2"] = "Cash Flow ($mm) — quarterly"
    ws["B2"].font = TITLE

    header_row = 5
    ws.cell(row=header_row, column=2, value="Line item").font = BOLD_WHITE
    ws.cell(row=header_row, column=2).fill = HEADER_FILL
    for i, lbl in enumerate(all_labels):
        c = ws.cell(row=header_row, column=3 + i, value=lbl)
        c.font = BOLD_WHITE
        c.fill = HEADER_FILL
        c.alignment = CENTER

    rows_spec = [
        ("Operating Cash Flow", "Operating Cash Flow", True),
        ("Capital Expenditure", "Capital Expenditure", False),
        ("Free Cash Flow", "Free Cash Flow", True),
        ("Stock-Based Comp", "Stock Based Compensation", False),
        ("D&A", "Depreciation And Amortization", False),
    ]
    n_hist = len(q_labels)
    FCF_ROW = 8
    r = 6
    for lbl, key, bold in rows_spec:
        ws.cell(row=r, column=2, value=lbl).font = BOLD_BLACK if bold else BLACK
        for i, p in enumerate(periods):
            v = p.get(key)
            c = ws.cell(row=r, column=3 + i, value=(v / 1e6) if v is not None else None)
            c.font = BLUE
            c.number_format = FMT_MONEY_MM
        r += 1

    # Forecast FCF = Rev × (fcf_sbc_margin + sbc_pct_rev)
    # Link revenue from Income Stmt (cross-sheet → GREEN)
    pnl = wb._pnl_coords                       # type: ignore[attr-defined]
    rev_row_q = 6
    for j in range(len(forecast_q)):
        is_col = get_column_letter(3 + n_hist + j)
        y = j + 1
        ramp = min(1.0, y / (4 * MARGIN_RAMP_YEARS))
        # approximate FCF at quarterly granularity: revenue × (ramped FCF-SBC margin + SBC%)
        formula = (
            f"='Income Stmt'!{is_col}{rev_row_q}*(({refs['ttm_fcf_sbc']}/{refs['ttm_revenue']})"
            f"+({refs['fcf_sbc_margin_target']}-{refs['ttm_fcf_sbc']}/{refs['ttm_revenue']})*{ramp}"
            f"+{refs['sbc_pct_rev']})"
        )
        c = ws.cell(row=FCF_ROW, column=3 + n_hist + j, value=formula)
        c.font = GREEN
        c.number_format = FMT_MONEY_MM


# ---------------------------------------------------------------------------
# Sheet: Valuation — 3 scenarios (Downside / Base / Upside)
# ---------------------------------------------------------------------------
def _build_valuation(wb: Workbook, fin: FinancialData, t: TargetResult) -> None:
    ws = wb.create_sheet("Valuation")
    refs = wb._target_refs                     # type: ignore[attr-defined]
    scen_refs = wb._scen_refs                  # type: ignore[attr-defined]
    pnl = wb._pnl_coords                       # type: ignore[attr-defined]

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 38
    for col in ["C", "D", "E"]:
        ws.column_dimensions[col].width = 17

    ws["B2"] = "Valuation — 3-scenario, EV/EBITDA × EV/(FCF-SBC) blend, WACC-discounted"
    ws["B2"].font = TITLE
    ws["B3"] = f"Method: value at end of Year {VALUATION_YEAR}, discount back {DISCOUNT_YEARS} years (target is 1-yr forward)."
    ws["B3"].font = ITALIC

    # Header row
    ws["C5"] = "Downside"
    ws["D5"] = "Base"
    ws["E5"] = "Upside"
    for col in ["C", "D", "E"]:
        c = ws[f"{col}5"]
        c.font = BOLD_WHITE
        c.fill = HEADER_FILL
        c.alignment = CENTER
    ws["D5"].fill = PatternFill("solid", start_color="2E5CB8")  # emphasize base

    # Emphasize Base column with a subtle tint on values below
    for rr in range(6, 40):
        ws.cell(row=rr, column=4).fill = BASE_COL_FILL

    # ---- Scenario-specific drivers (derived from base × offsets) ----
    # Row 6: scenario label row already done. Start drivers at row 7.
    r = 7
    ws.cell(row=r, column=2, value="— Scenario drivers —").font = BOLD_BLACK
    r += 1
    # rev growth Y1
    ws.cell(row=r, column=2, value="Revenue growth Y1").font = BLACK
    for scen, col in SCEN_COLS.items():
        c = ws.cell(row=r, column=ord(col) - ord("A") + 1,
                    value=f"={refs['rev_growth_y1']}*{scen_refs[scen]['rev_growth_y1_mult']}")
        c.font = BLACK
        c.number_format = FMT_PCT
    REVGR_Y1_ROW = r
    r += 1
    # rev growth terminal
    ws.cell(row=r, column=2, value="Revenue growth terminal").font = BLACK
    for scen, col in SCEN_COLS.items():
        c = ws.cell(row=r, column=ord(col) - ord("A") + 1,
                    value=f"={refs['rev_growth_terminal']}*{scen_refs[scen]['rev_growth_terminal_mult']}")
        c.font = BLACK
        c.number_format = FMT_PCT
    REVGR_T_ROW = r
    r += 1
    # EBITDA margin target
    ws.cell(row=r, column=2, value="EBITDA margin (Y3 target)").font = BLACK
    for scen, col in SCEN_COLS.items():
        c = ws.cell(row=r, column=ord(col) - ord("A") + 1,
                    value=f"={refs['ebitda_margin_target']}+{scen_refs[scen]['ebitda_margin_delta']}")
        c.font = BLACK
        c.number_format = FMT_PCT
    EB_MGN_ROW = r
    r += 1
    # FCF-SBC margin target
    ws.cell(row=r, column=2, value="FCF − SBC margin (Y3 target)").font = BLACK
    for scen, col in SCEN_COLS.items():
        c = ws.cell(row=r, column=ord(col) - ord("A") + 1,
                    value=f"={refs['fcf_sbc_margin_target']}+{scen_refs[scen]['fcf_sbc_margin_delta']}")
        c.font = BLACK
        c.number_format = FMT_PCT
    FCF_MGN_ROW = r
    r += 1
    # EV/EBITDA mult
    ws.cell(row=r, column=2, value="EV/EBITDA mult (NTM+3)").font = BLACK
    for scen, col in SCEN_COLS.items():
        c = ws.cell(row=r, column=ord(col) - ord("A") + 1,
                    value=f"={refs['ev_ebitda_multiple']}*{scen_refs[scen]['ev_ebitda_multiple_mult']}")
        c.font = BLACK
        c.number_format = FMT_MULT
    EV_EB_MULT_ROW = r
    r += 1
    # EV/(FCF-SBC) mult
    ws.cell(row=r, column=2, value="EV/(FCF-SBC) mult (NTM+3)").font = BLACK
    for scen, col in SCEN_COLS.items():
        c = ws.cell(row=r, column=ord(col) - ord("A") + 1,
                    value=f"={refs['ev_fcf_sbc_multiple']}*{scen_refs[scen]['ev_fcf_sbc_multiple_mult']}")
        c.font = BLACK
        c.number_format = FMT_MULT
    EV_FCF_MULT_ROW = r
    r += 1
    # WACC
    ws.cell(row=r, column=2, value="WACC").font = BLACK
    for scen, col in SCEN_COLS.items():
        c = ws.cell(row=r, column=ord(col) - ord("A") + 1,
                    value=f"={scen_refs[scen]['discount_rate']}")
        c.font = BLACK
        c.number_format = FMT_PCT
    WACC_ROW = r
    r += 2

    # ---- Y3 forecast by scenario ----
    # To compute Y3 revenue with a linearly-decaying growth, we apply
    # product of (1 + g_y) for y = 1..3 where g_y = g1 + (gT - g1)*(y-1)/(N-1)
    # With FORECAST_YEARS=5, weights for y=1,2,3 are 0, 1/4, 2/4.
    ws.cell(row=r, column=2, value="— Year 3 forecast —").font = BOLD_BLACK
    r += 1
    # Revenue Y3
    ws.cell(row=r, column=2, value="Revenue Y3 ($mm)").font = BOLD_BLACK
    for scen, col in SCEN_COLS.items():
        g1_c = f"{col}{REVGR_Y1_ROW}"
        gT_c = f"{col}{REVGR_T_ROW}"
        # factors for y=1,2,3
        # y1: (1 + g1)
        # y2: (1 + g1 + (gT-g1)/4)
        # y3: (1 + g1 + 2*(gT-g1)/4)
        formula = (
            f"={refs['ttm_revenue']}"
            f"*(1+{g1_c})"
            f"*(1+{g1_c}+({gT_c}-{g1_c})/4)"
            f"*(1+{g1_c}+2*({gT_c}-{g1_c})/4)"
        )
        c = ws.cell(row=r, column=ord(col) - ord("A") + 1, value=formula)
        c.font = BLACK
        c.number_format = FMT_MONEY_MM
    REV_Y3_ROW = r
    r += 1
    # EBITDA margin Y3 = ramp ≤ 1.0 at y=3 with MARGIN_RAMP_YEARS=3 → fully at target
    ws.cell(row=r, column=2, value="EBITDA margin Y3").font = BLACK
    for scen, col in SCEN_COLS.items():
        c = ws.cell(row=r, column=ord(col) - ord("A") + 1, value=f"={col}{EB_MGN_ROW}")
        c.font = BLACK
        c.number_format = FMT_PCT
    r += 1
    # EBITDA Y3
    ws.cell(row=r, column=2, value="EBITDA Y3 ($mm)").font = BOLD_BLACK
    for scen, col in SCEN_COLS.items():
        c = ws.cell(row=r, column=ord(col) - ord("A") + 1,
                    value=f"={col}{REV_Y3_ROW}*{col}{EB_MGN_ROW}")
        c.font = BLACK
        c.number_format = FMT_MONEY_MM
    EBITDA_Y3_ROW = r
    r += 1
    # FCF-SBC margin Y3 = ramp ≤ 1.0 → fully at target
    ws.cell(row=r, column=2, value="FCF − SBC margin Y3").font = BLACK
    for scen, col in SCEN_COLS.items():
        c = ws.cell(row=r, column=ord(col) - ord("A") + 1, value=f"={col}{FCF_MGN_ROW}")
        c.font = BLACK
        c.number_format = FMT_PCT
    r += 1
    # FCF-SBC Y3
    ws.cell(row=r, column=2, value="FCF − SBC Y3 ($mm)").font = BOLD_BLACK
    for scen, col in SCEN_COLS.items():
        c = ws.cell(row=r, column=ord(col) - ord("A") + 1,
                    value=f"={col}{REV_Y3_ROW}*{col}{FCF_MGN_ROW}")
        c.font = BLACK
        c.number_format = FMT_MONEY_MM
    FCFSBC_Y3_ROW = r
    r += 2

    # ---- Terminal EV ----
    ws.cell(row=r, column=2, value="— Terminal enterprise value (end Y3) —").font = BOLD_BLACK
    r += 1
    ws.cell(row=r, column=2, value="EV via EBITDA × mult").font = BLACK
    for scen, col in SCEN_COLS.items():
        c = ws.cell(row=r, column=ord(col) - ord("A") + 1,
                    value=f"={col}{EBITDA_Y3_ROW}*{col}{EV_EB_MULT_ROW}")
        c.font = BLACK
        c.number_format = FMT_MONEY_MM
    EV_EB_ROW = r
    r += 1
    ws.cell(row=r, column=2, value="EV via (FCF-SBC) × mult").font = BLACK
    for scen, col in SCEN_COLS.items():
        c = ws.cell(row=r, column=ord(col) - ord("A") + 1,
                    value=f"={col}{FCFSBC_Y3_ROW}*{col}{EV_FCF_MULT_ROW}")
        c.font = BLACK
        c.number_format = FMT_MONEY_MM
    EV_FCF_ROW = r
    r += 1
    ws.cell(row=r, column=2, value="Terminal EV (50/50 blend)").font = BOLD_BLACK
    for scen, col in SCEN_COLS.items():
        c = ws.cell(row=r, column=ord(col) - ord("A") + 1,
                    value=f"=({col}{EV_EB_ROW}+{col}{EV_FCF_ROW})/2")
        c.font = BLACK
        c.number_format = FMT_MONEY_MM
    TERM_EV_ROW = r
    r += 2

    # ---- PV of terminal EV ----
    ws.cell(row=r, column=2, value="— PV & equity per share —").font = BOLD_BLACK
    r += 1
    ws.cell(row=r, column=2, value=f"PV factor = 1 / (1+WACC)^{DISCOUNT_YEARS}").font = BLACK
    for scen, col in SCEN_COLS.items():
        c = ws.cell(row=r, column=ord(col) - ord("A") + 1,
                    value=f"=1/(1+{col}{WACC_ROW})^{DISCOUNT_YEARS}")
        c.font = BLACK
        c.number_format = "0.0000"
    PV_FACTOR_ROW = r
    r += 1
    ws.cell(row=r, column=2, value="PV of terminal EV").font = BOLD_BLACK
    for scen, col in SCEN_COLS.items():
        c = ws.cell(row=r, column=ord(col) - ord("A") + 1,
                    value=f"={col}{TERM_EV_ROW}*{col}{PV_FACTOR_ROW}")
        c.font = BLACK
        c.number_format = FMT_MONEY_MM
    PV_EV_ROW = r
    r += 1
    ws.cell(row=r, column=2, value="− Net debt").font = BLACK
    for scen, col in SCEN_COLS.items():
        c = ws.cell(row=r, column=ord(col) - ord("A") + 1, value=f"={refs['net_debt']}")
        c.font = GREEN
        c.number_format = FMT_MONEY_MM
    ND_ROW = r
    r += 1
    ws.cell(row=r, column=2, value="→ Equity value").font = BOLD_BLACK
    for scen, col in SCEN_COLS.items():
        c = ws.cell(row=r, column=ord(col) - ord("A") + 1,
                    value=f"={col}{PV_EV_ROW}-{col}{ND_ROW}")
        c.font = BLACK
        c.number_format = FMT_MONEY_MM
    EQ_ROW = r
    r += 1
    ws.cell(row=r, column=2,
            value=f"÷ Diluted shares (adj +{DISCOUNT_YEARS}y dilution)").font = BLACK
    for scen, col in SCEN_COLS.items():
        c = ws.cell(row=r, column=ord(col) - ord("A") + 1,
                    value=f"={refs['shares_0']}*(1+{refs['share_change_pct']})^{VALUATION_YEAR - DISCOUNT_YEARS}")
        c.font = BLACK
        c.number_format = FMT_SHARES
    SHARES_ROW = r
    r += 1
    ws.cell(row=r, column=2, value="→ Price per share").font = BOLD_BLACK
    ws.cell(row=r, column=2).fill = TOTAL_FILL
    for scen, col in SCEN_COLS.items():
        c = ws.cell(row=r, column=ord(col) - ord("A") + 1,
                    value=f"=IFERROR({col}{EQ_ROW}/{col}{SHARES_ROW},0)")
        c.font = BOLD_BLACK
        c.fill = TOTAL_FILL
        c.number_format = FMT_PRICE
    PRICE_ROW = r
    r += 1
    ws.cell(row=r, column=2, value="Upside vs current").font = BLACK
    for scen, col in SCEN_COLS.items():
        c = ws.cell(row=r, column=ord(col) - ord("A") + 1,
                    value=f"=IFERROR({col}{PRICE_ROW}/{refs['price_0']}-1,0)")
        c.font = BLACK
        c.number_format = FMT_PCT
    r += 2

    # Side-by-side method comparison (base only — quick sanity)
    ws.cell(row=r, column=2, value="— Base scenario: method comparison —").font = BOLD_BLACK
    r += 1
    ws.cell(row=r, column=2, value="Price via EV/EBITDA only").font = BLACK
    c = ws.cell(row=r, column=4,
                value=f"=IFERROR((D{EV_EB_ROW}*D{PV_FACTOR_ROW}-{refs['net_debt']})/D{SHARES_ROW},0)")
    c.font = BLACK
    c.number_format = FMT_PRICE
    r += 1
    ws.cell(row=r, column=2, value="Price via EV/(FCF-SBC) only").font = BLACK
    c = ws.cell(row=r, column=4,
                value=f"=IFERROR((D{EV_FCF_ROW}*D{PV_FACTOR_ROW}-{refs['net_debt']})/D{SHARES_ROW},0)")
    c.font = BLACK
    c.number_format = FMT_PRICE


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    import sys
    args = sys.argv[1:]
    tk = (args[0] if args else "APP").upper()

    # Parse optional kv args: horizon_months=24 + out_dir positional
    horizon_months = 12
    out_dir = Path(
        "/sessions/fervent-charming-johnson/mnt/Agent System for finding Stocks/stock-radar/out"
    )
    for a in args[1:]:
        if "=" in a:
            k, v = a.split("=", 1)
            if k == "horizon_months":
                try:
                    horizon_months = int(float(v))
                except ValueError:
                    pass
        else:
            out_dir = Path(a)

    try:
        fin = fetch_financials(tk)
    except EarningsFetchError as e:
        print(f"[FAIL] {e}")
        sys.exit(1)
    target = build_target(fin, horizon_months=horizon_months)
    path = export_model(fin, target, out_dir / f"{tk}_model.xlsx")
    print(f"[OK] wrote {path}")
    print(f"     Current: ${target.current_price:.2f}   "
          f"Low: ${target.low:.2f}   Base: ${target.base:.2f}   "
          f"High: ${target.high:.2f}")
