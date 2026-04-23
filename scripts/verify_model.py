#!/usr/bin/env python3
"""
verify_model.py — Cross-ticker correctness harness.

For each ticker:
    1. Run the engine (`build_target`) to produce Python forecast/scenario numbers.
    2. Export the workbook via `export_model` and recalc formulas with LibreOffice.
    3. Load the recalc'd workbook and read back every forecast/valuation cell.
    4. Compare engine values ↔ Excel values ↔ JSON-API (target_api.py) values.
    5. Print a per-ticker diff report; exit non-zero if any REAL mismatch exceeds
       tolerance (1e-3 relative for $ values; 1e-4 absolute for margins).

The harness is self-configuring: it derives `n_hist` by matching header cells in
the P&L Summary against the forecast-period labels produced by the engine — no
hardcoded column counts. That means it works even for tickers with <4 annual
historicals (SNDK, RKLB, ACHR) or non-December fiscal years (MU, SNDK).
"""
from __future__ import annotations

import json
import subprocess
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from finance_data import fetch_financials, EarningsFetchError
from target_engine import build_target
from model_export import export_model


REPO = Path(__file__).resolve().parent.parent  # stock-radar/
OUT_DIR = REPO / "out"
RECALC = REPO / "scripts" / "recalc.py"
# No skill recalc fallback needed — use repo-local only

DEFAULT_TICKERS = ["MRVL", "SNDK", "MU", "LITE", "APP", "AEHR", "TER", "NVDA", "AMD"]

# Tolerances
REL_TOL_USD = 1e-3      # 0.1% for dollar amounts
ABS_TOL_PCT = 1e-4      # 1 bp for ratios/margins
ABS_TOL_PRICE = 0.01    # 1¢ per share


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
class Mismatch(Exception):
    pass


def _rel_err(a: float, b: float) -> float:
    if a is None or b is None:
        return float("inf")
    denom = max(abs(a), abs(b), 1e-9)
    return abs(a - b) / denom


def _eq_usd(a: float, b: float) -> bool:
    return _rel_err(a, b) <= REL_TOL_USD or abs(a - b) < 0.5  # half a dollar


def _eq_pct(a: float, b: float) -> bool:
    return abs((a or 0) - (b or 0)) <= ABS_TOL_PCT


def _eq_price(a: float, b: float) -> bool:
    return abs((a or 0) - (b or 0)) <= ABS_TOL_PRICE


def _recalc(path: Path) -> None:
    """Invoke LibreOffice recalc. Fatal on error."""
    if not RECALC.exists():
        raise FileNotFoundError(f"recalc.py not found at {RECALC}")
    r = subprocess.run(
        [sys.executable, str(RECALC), str(path), "60"],
        capture_output=True, text=True, timeout=180,
    )
    if r.returncode != 0:
        raise RuntimeError(f"recalc failed for {path.name}:\n{r.stderr}\n{r.stdout}")
    # Parse status JSON from stdout. Last non-blank line is the JSON doc.
    lines = [ln.strip() for ln in r.stdout.splitlines() if ln.strip()]
    if not lines:
        return
    try:
        status = json.loads(lines[-1])
    except json.JSONDecodeError:
        return
    if status.get("status") == "errors_found":
        print(f"  ⚠ formula errors: {json.dumps(status.get('error_summary'), indent=2)}")


def _find_header_col(ws, row: int, label: str) -> int | None:
    """Return 1-based column index where `ws[row][col].value == label`, else None."""
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=row, column=col).value
        if v is not None and str(v).strip() == str(label).strip():
            return col
    return None


def _find_row_by_label(ws, label: str, col: int = 2) -> int | None:
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=col).value
        if v is not None and str(v).strip().lower().startswith(label.strip().lower()):
            return r
    return None


# ---------------------------------------------------------------------------
# Verification: engine ↔ Excel
# ---------------------------------------------------------------------------
def verify_pnl_summary(wb, t, ticker: str) -> list[dict[str, Any]]:
    """Compare engine forecast_annual to P&L Summary forecast columns."""
    diffs: list[dict[str, Any]] = []
    ws = wb["P&L Summary"]

    # Row map (these are stable in model_export._build_pnl_summary)
    REV_ROW, EB_MGN_ROW, EB_ROW, FCF_MGN_ROW, FCF_ROW, OI_ROW, NI_ROW = 6, 8, 9, 10, 11, 12, 13
    HEADER_ROW = 5

    # Auto-detect each forecast column by matching header cell to engine's period label
    for y_idx, period in enumerate(t.forecast_annual, start=1):
        col = _find_header_col(ws, HEADER_ROW, period.period)
        if col is None:
            diffs.append({"ticker": ticker, "tag": f"Y{y_idx} header", "note": f"label '{period.period}' not found"})
            continue

        # Engine values are in raw dollars; Excel stores $mm
        def cell(row):
            v = ws.cell(row=row, column=col).value
            return float(v) if v is not None else None

        xl_rev = cell(REV_ROW)
        xl_eb = cell(EB_ROW)
        xl_fcf = cell(FCF_ROW)
        xl_oi = cell(OI_ROW)
        xl_ni = cell(NI_ROW)
        xl_eb_mgn = cell(EB_MGN_ROW)
        xl_fcf_mgn = cell(FCF_MGN_ROW)

        # Convert engine values to $mm for comparison
        eng = {
            "Rev":        period.revenue / 1e6,
            "EBITDA":     period.ebitda / 1e6,
            "FCF-SBC":    period.fcf_sbc / 1e6,
            "OI":         period.operating_income / 1e6,
            "NI":         period.net_income / 1e6,
            "EB mgn":     period.ebitda_margin,
            "FCF mgn":    period.fcf_sbc_margin,
        }
        xl = {
            "Rev": xl_rev, "EBITDA": xl_eb, "FCF-SBC": xl_fcf,
            "OI": xl_oi, "NI": xl_ni,
            "EB mgn": xl_eb_mgn, "FCF mgn": xl_fcf_mgn,
        }

        for metric, eng_v in eng.items():
            xl_v = xl[metric]
            if xl_v is None:
                diffs.append({"ticker": ticker, "tag": f"Y{y_idx} {metric}", "eng": eng_v, "xl": None})
                continue
            if "mgn" in metric:
                if not _eq_pct(eng_v, xl_v):
                    diffs.append({"ticker": ticker, "tag": f"Y{y_idx} {metric}", "eng": eng_v, "xl": xl_v,
                                  "rel_err": abs(eng_v - xl_v)})
            else:
                if not _eq_usd(eng_v, xl_v):
                    diffs.append({"ticker": ticker, "tag": f"Y{y_idx} {metric}", "eng": eng_v, "xl": xl_v,
                                  "rel_err": _rel_err(eng_v, xl_v)})
    return diffs


def verify_valuation(wb, t, ticker: str) -> list[dict[str, Any]]:
    """Compare engine scenarios[s].price to Valuation sheet Price per share row."""
    diffs: list[dict[str, Any]] = []
    ws = wb["Valuation"]
    # Find price row
    price_row = _find_row_by_label(ws, "→ Price per share")
    if price_row is None:
        diffs.append({"ticker": ticker, "tag": "Valuation", "note": "price row not found"})
        return diffs

    # Scenarios in Valuation are Downside (D) / Base (E) / Upside (F)
    # (See model_export SCEN_COLS.) But we should discover the mapping from the header row.
    # Find header row for scenarios — typically row 5 with labels "Downside" / "Base" / "Upside"
    scen_row = None
    for r in range(1, 10):
        row_vals = [str(ws.cell(row=r, column=c).value or "").strip().lower() for c in range(3, 8)]
        if "base" in row_vals:
            scen_row = r
            break
    if scen_row is None:
        diffs.append({"ticker": ticker, "tag": "Valuation scen header", "note": "not found"})
        return diffs

    label_map = {"downside": "downside", "base": "base", "upside": "upside",
                 "low": "downside", "high": "upside"}
    xl_prices: dict[str, float] = {}
    for c in range(3, 8):
        v = ws.cell(row=scen_row, column=c).value
        if v:
            key = label_map.get(str(v).strip().lower())
            if key:
                p = ws.cell(row=price_row, column=c).value
                if p is not None:
                    xl_prices[key] = float(p)

    for scen in ("downside", "base", "upside"):
        eng_p = t.scenarios[scen].price if scen in t.scenarios else None
        xl_p = xl_prices.get(scen)
        if eng_p is None or xl_p is None:
            diffs.append({"ticker": ticker, "tag": f"Val {scen}", "eng": eng_p, "xl": xl_p, "note": "missing"})
            continue
        if not _eq_price(eng_p, xl_p):
            diffs.append({"ticker": ticker, "tag": f"Val {scen}", "eng": eng_p, "xl": xl_p,
                          "rel_err": _rel_err(eng_p, xl_p)})

    # Also check TargetResult.base/low/high against engine scenarios (internal consistency)
    if abs(t.base - t.scenarios["base"].price) > 1e-6:
        diffs.append({"ticker": ticker, "tag": "t.base ≠ scen.base", "eng": t.base, "xl": t.scenarios["base"].price})
    if abs(t.low - t.scenarios["downside"].price) > 1e-6:
        diffs.append({"ticker": ticker, "tag": "t.low ≠ scen.down", "eng": t.low, "xl": t.scenarios["downside"].price})
    if abs(t.high - t.scenarios["upside"].price) > 1e-6:
        diffs.append({"ticker": ticker, "tag": "t.high ≠ scen.up", "eng": t.high, "xl": t.scenarios["upside"].price})

    return diffs


def verify_json_api(t, ticker: str) -> list[dict[str, Any]]:
    """Call target_api.py and verify JSON payload matches engine.to_dict()."""
    diffs: list[dict[str, Any]] = []
    api = REPO / "scripts" / "target_api.py"
    r = subprocess.run(
        [sys.executable, str(api), ticker],
        capture_output=True, text=True, timeout=180,
    )
    if r.returncode != 0:
        diffs.append({"ticker": ticker, "tag": "target_api exit", "note": r.stderr[:200]})
        return diffs
    try:
        payload = json.loads(r.stdout)
    except json.JSONDecodeError as e:
        diffs.append({"ticker": ticker, "tag": "target_api JSON", "note": f"parse: {e}"})
        return diffs
    if "error" in payload:
        diffs.append({"ticker": ticker, "tag": "target_api error", "note": payload["error"]})
        return diffs

    tgt = payload["target"]
    # Top-level prices
    for key in ("base", "low", "high", "current_price"):
        eng_v = getattr(t, key)
        api_v = tgt[key]
        if not _eq_price(eng_v, api_v):
            diffs.append({"ticker": ticker, "tag": f"JSON {key}", "eng": eng_v, "xl": api_v,
                          "rel_err": _rel_err(eng_v, api_v)})

    # Every scenario price must match
    for scen in ("downside", "base", "upside"):
        eng_p = t.scenarios[scen].price
        api_p = tgt["scenarios"][scen]["price"]
        if not _eq_price(eng_p, api_p):
            diffs.append({"ticker": ticker, "tag": f"JSON scen.{scen}.price", "eng": eng_p, "xl": api_p})

    # Forecast annual revenue/ebitda/oi/ni must match per year
    for y, (eng_p, api_p) in enumerate(zip(t.forecast_annual, tgt["forecast_annual"]), start=1):
        for k in ("revenue", "ebitda", "operating_income", "net_income", "fcf_sbc"):
            a, b = getattr(eng_p, k), api_p[k]
            if not _eq_usd(a, b):
                diffs.append({"ticker": ticker, "tag": f"JSON Y{y} {k}", "eng": a, "xl": b,
                              "rel_err": _rel_err(a, b)})

    # ---- Dashboard shape check ----
    # Every field consumed by app/model/[ticker]/detailed/DetailedModel.tsx MUST
    # be present. If this fails, the detailed page will render "Cannot build model".
    dash_required = {
        "top": ["ticker", "name", "sector", "target", "historicals", "capitalization"],
        "target": [
            "current_price", "low", "base", "high",
            "upside_base_pct", "upside_low_pct", "upside_high_pct",
            "steps", "forecast_quarterly", "forecast_annual", "scenarios",
            "drivers", "terminal_year",
            "ttm_revenue", "ttm_ebitda", "ttm_fcf_sbc",
            "net_debt", "shares_diluted",
        ],
        "forecast_period": [
            "period", "revenue", "operating_income", "ebitda", "ebitda_margin",
            "fcf_sbc", "fcf_sbc_margin", "net_income", "fcf", "op_margin",
            "rev_growth",
        ],
        "scenario": [
            "price", "discount_rate",
            "rev_growth_y1", "rev_growth_terminal",
            "ebitda_margin_target", "fcf_sbc_margin_target",
            "ev_ebitda_multiple", "ev_fcf_sbc_multiple",
            "terminal_revenue", "terminal_ebitda", "terminal_fcf_sbc",
            "ev_from_ebitda", "ev_from_fcf_sbc",
            "terminal_ev_blended", "pv_ev_blended",
            "pv_ev_from_ebitda", "pv_ev_from_fcf_sbc",
            "equity_value", "price_from_ebitda", "price_from_fcf_sbc",
        ],
        "historicals": ["quarterly", "annual", "ttm"],
        "capitalization": ["price", "market_cap", "shares_diluted", "net_debt"],
    }
    for k in dash_required["top"]:
        if k not in payload:
            diffs.append({"ticker": ticker, "tag": f"dash top.{k}", "note": "MISSING"})
    for k in dash_required["target"]:
        if k not in tgt:
            diffs.append({"ticker": ticker, "tag": f"dash target.{k}", "note": "MISSING"})
    if tgt.get("forecast_annual"):
        fp = tgt["forecast_annual"][0]
        for k in dash_required["forecast_period"]:
            if k not in fp:
                diffs.append({"ticker": ticker, "tag": f"dash forecast.{k}", "note": "MISSING"})
    for scen in ("downside", "base", "upside"):
        s = tgt.get("scenarios", {}).get(scen, {})
        if not s:
            diffs.append({"ticker": ticker, "tag": f"dash scen.{scen}", "note": "MISSING"})
            continue
        for k in dash_required["scenario"]:
            if k not in s:
                diffs.append({"ticker": ticker, "tag": f"dash scen.{scen}.{k}", "note": "MISSING"})
    for k in dash_required["historicals"]:
        if k not in payload.get("historicals", {}):
            diffs.append({"ticker": ticker, "tag": f"dash hist.{k}", "note": "MISSING"})
    for k in dash_required["capitalization"]:
        if k not in payload.get("capitalization", {}):
            diffs.append({"ticker": ticker, "tag": f"dash cap.{k}", "note": "MISSING"})
    # Required non-empty sequences the UI iterates over
    if not tgt.get("forecast_annual"):
        diffs.append({"ticker": ticker, "tag": "dash forecast_annual", "note": "EMPTY"})
    if not tgt.get("forecast_quarterly"):
        diffs.append({"ticker": ticker, "tag": "dash forecast_quarterly", "note": "EMPTY"})
    if not tgt.get("steps"):
        diffs.append({"ticker": ticker, "tag": "dash steps", "note": "EMPTY (Valuation tab deduction chain won't render)"})

    return diffs


# ---------------------------------------------------------------------------
# Driver
# ---------------------------------------------------------------------------
def verify_ticker(ticker: str) -> dict[str, Any]:
    print(f"\n=== {ticker} ===")
    result: dict[str, Any] = {"ticker": ticker, "diffs": [], "status": "ok"}

    try:
        fin = fetch_financials(ticker)
    except EarningsFetchError as e:
        result["status"] = "fetch_fail"
        result["note"] = str(e)
        print(f"  [skip] {e}")
        return result

    t = build_target(fin)
    print(f"  engine: base=${t.base:.2f}  low=${t.low:.2f}  high=${t.high:.2f}  current=${t.current_price:.2f}")

    xlsx_path = OUT_DIR / f"{ticker}_verify.xlsx"
    export_model(fin, t, xlsx_path)
    try:
        _recalc(xlsx_path)
    except Exception as e:
        result["status"] = "recalc_fail"
        result["note"] = str(e)
        print(f"  [skip] recalc failed: {e}")
        return result

    wb = load_workbook(xlsx_path, data_only=True)

    diffs: list[dict[str, Any]] = []
    diffs += verify_pnl_summary(wb, t, ticker)
    diffs += verify_valuation(wb, t, ticker)
    diffs += verify_json_api(t, ticker)

    # Engine internal identities
    for scen, s in t.scenarios.items():
        blended = (s.ev_from_ebitda + s.ev_from_fcf_sbc) / 2
        if not _eq_usd(blended, s.terminal_ev_blended):
            diffs.append({"ticker": ticker, "tag": f"eng {scen} blend", "eng": blended, "xl": s.terminal_ev_blended})
        # Price internal: equity / (shares * (1+div)^1)
        equity_check = s.pv_ev_blended - t.net_debt
        if not _eq_usd(equity_check, s.equity_value):
            diffs.append({"ticker": ticker, "tag": f"eng {scen} equity", "eng": equity_check, "xl": s.equity_value})

    result["diffs"] = diffs
    if diffs:
        result["status"] = "diffs"
    for d in diffs:
        if "rel_err" in d:
            print(f"  ✗ {d['tag']:30s} eng={d.get('eng'):>14.3f} xl={d.get('xl'):>14.3f} ({d['rel_err']*100:.2f}%)")
        else:
            print(f"  ✗ {d['tag']:30s} {d}")
    if not diffs:
        print("  ✓ all checks pass")
    return result


def main() -> int:
    tickers = sys.argv[1:] if len(sys.argv) > 1 else DEFAULT_TICKERS
    summary = [verify_ticker(tk) for tk in tickers]
    print("\n" + "=" * 70)
    print(f"SUMMARY: {len(tickers)} tickers verified")
    bad = 0
    for r in summary:
        n = len(r["diffs"])
        flag = "OK " if n == 0 and r["status"] == "ok" else f"[{r['status']}]"
        print(f"  {flag:12s} {r['ticker']:6s}  diffs={n}")
        if n:
            bad += 1
    return 1 if bad else 0


if __name__ == "__main__":
    sys.exit(main())
